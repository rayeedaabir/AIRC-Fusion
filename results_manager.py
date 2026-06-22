"""
results_manager.py  (NEW - added for the Cureus/CJCS revision)
------------------------------------------------------------------
Single place that collects the results of every (model x dataset) run and
writes them out in paper-ready form, so you never copy-paste a terminal
summary again.

For each run it stores: image counts, the mean of every metric, the 95%
confidence interval (computed exactly like calculate_confidence_intervals.py
-> scipy t-interval), and min/max. It then writes:

  results/master_metrics_long.csv      one row per (dataset, model), every metric
  results/AIRC-Fusion_Results.xlsx     formatted workbook with:
      - "All_runs"        : the long table above
      - "Core_<dataset>"  : paper Table-6 layout (metrics x models) per dataset
      - "Temporal_Score"  : paper Table-7 layout (datasets x models)
      - "FPS" / "SSIM"    : quick cross-dataset pivots

Nothing here touches the fusion math or the metric math; it only consumes
the dictionaries your DatasetProcessor already returns.
"""

import os
import numpy as np
import pandas as pd
from scipy import stats

# Canonical metric keys -> (pretty label, per-frame column candidates, higher_is_better)
# The per-frame column names differ slightly between the eval metrics and the
# temporal metrics, so we list every candidate name we might encounter.
METRIC_SPECS = [
    ("proc_time_ms",  "Proc Time (ms)",   ["proc_time"],                 False, 1000.0),
    ("fps",           "FPS",              ["fps"],                       True,  1.0),
    ("ssim_accuracy", "SSIM",             ["ssim_accuracy"],             True,  1.0),
    ("entropy",       "Entropy",          ["entropy"],                   True,  1.0),
    ("mutual_info_src", "Mutual Info",    ["mutual_info_src"],           True,  1.0),
    ("std_dev",       "Std Dev",          ["std_dev"],                   True,  1.0),
    ("edge_sum",      "Edge Sum",         ["edge_sum"],                  True,  1.0),
    ("noise",         "Noise",            ["noise"],                     False, 1.0),
    ("wavelet_std_cA","Wavelet cA",       ["wavelet_std_cA"],            True,  1.0),
    ("wavelet_std_cH","Wavelet cH",       ["wavelet_std_cH"],            True,  1.0),
    ("wavelet_std_cV","Wavelet cV",       ["wavelet_std_cV"],            True,  1.0),
    ("wavelet_std_cD","Wavelet cD",       ["wavelet_std_cD"],            True,  1.0),
    ("matd_fused",    "MATD Fused",       ["MATD_fused", "temporal_diff_fused"], False, 1.0),
    ("temporal_instability_score", "Temporal Instability Score",
                       ["Temporal_Instability_Score", "temporal_instability_score"], False, 1.0),
]


def _ci95(series):
    """Mean and 95% CI using the t-distribution (matches calculate_confidence_intervals.py)."""
    data = pd.to_numeric(series, errors="coerce").dropna()
    if len(data) < 2:
        m = float(data.mean()) if len(data) == 1 else np.nan
        return m, np.nan, np.nan, (float(data.min()) if len(data) else np.nan), (float(data.max()) if len(data) else np.nan)
    mean = float(np.mean(data))
    sem = float(stats.sem(data))
    lo, hi = stats.t.interval(0.95, df=len(data) - 1, loc=mean, scale=sem) if sem > 0 else (mean, mean)
    return mean, float(lo), float(hi), float(np.min(data)), float(np.max(data))


class ResultsManager:
    def __init__(self, out_dir):
        self.out_dir = out_dir
        os.makedirs(out_dir, exist_ok=True)
        self.rows = {}          # (dataset, model) -> row dict; ACCUMULATES across runs
        self._frames = {}       # (dataset, model) -> per-frame DataFrame (for re-pivots)
        # Resume support: load previously-saved rows so dataset-by-dataset runs
        # build up the SAME workbook instead of overwriting it.
        _csv = os.path.join(out_dir, "master_metrics_long.csv")
        if os.path.exists(_csv):
            try:
                for _, r in pd.read_csv(_csv).iterrows():
                    self.rows[(r["dataset"], r["model"])] = r.to_dict()
                print(f"[results_manager] resumed {len(self.rows)} existing rows from {_csv}")
            except Exception as e:
                print(f"[results_manager] could not load existing CSV: {e}")

    def add_run(self, dataset_name, model_name, aggregated):
        """aggregated = the dict returned by DatasetProcessor.process_dataset_directory."""
        row = {
            "dataset": dataset_name,
            "model": model_name,
            "image_count": aggregated.get("image_count", np.nan),
            "processed_successfully": aggregated.get("processed_successfully", np.nan),
        }

        per_frame = aggregated.get("processed_pairs_details", None)
        pf_df = pd.DataFrame(per_frame) if per_frame else pd.DataFrame()
        self._frames[(dataset_name, model_name)] = pf_df

        for key, label, candidates, higher, scale in METRIC_SPECS:
            col = next((c for c in candidates if c in pf_df.columns), None)
            if col is not None and len(pf_df):
                mean, lo, hi, mn, mx = _ci95(pf_df[col] * scale)
            else:
                # Fallback: use the aggregated avg_* the processor already computed.
                avg = aggregated.get(f"avg_{candidates[-1]}", np.nan)
                mean = (avg * scale) if (avg is not None and not _isnan(avg)) else np.nan
                lo = hi = mn = mx = np.nan
            row[f"{key}_mean"] = mean
            row[f"{key}_ci_low"] = lo
            row[f"{key}_ci_high"] = hi
            row[f"{key}_min"] = mn
            row[f"{key}_max"] = mx
        self.rows[(dataset_name, model_name)] = row
        return row

    # ---------- exports ----------
    def long_dataframe(self):
        return pd.DataFrame(list(self.rows.values()))

    def done(self, dataset_name, model_name):
        """True if this (dataset, model) already has results (for resume/skip)."""
        return (dataset_name, model_name) in self.rows

    def save_csv(self):
        path = os.path.join(self.out_dir, "master_metrics_long.csv")
        self.long_dataframe().to_csv(path, index=False)
        return path

    def _fmt(self, mean, lo, hi, dec=3):
        if mean is None or _isnan(mean):
            return ""
        if lo is None or _isnan(lo):
            return f"{mean:.{dec}f}"
        return f"{mean:.{dec}f} ({lo:.{dec}f}, {hi:.{dec}f})"

    def save_excel(self):
        df = self.long_dataframe()
        path = os.path.join(self.out_dir, "AIRC-Fusion_Results.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as xw:
            df.to_csv  # noqa  (keep import warm)
            df.to_excel(xw, sheet_name="All_runs", index=False)

            # Paper Table-6 style: one sheet per dataset, metrics (rows) x models (cols),
            # each cell = "mean (ci_low, ci_high)".
            for ds in df["dataset"].unique():
                sub = df[df["dataset"] == ds]
                table = {}
                for _, r in sub.iterrows():
                    col = {}
                    for key, label, *_ in METRIC_SPECS:
                        col[label] = self._fmt(r[f"{key}_mean"], r[f"{key}_ci_low"], r[f"{key}_ci_high"])
                    table[r["model"]] = col
                tdf = pd.DataFrame(table)
                tdf.index.name = "Metric"
                sheet = ("Core_" + str(ds))[:31]
                tdf.to_excel(xw, sheet_name=sheet)

            # Paper Table-7 style: Temporal Instability Score, datasets (rows) x models (cols).
            pivot_tis = df.pivot_table(index="dataset", columns="model",
                                       values="temporal_instability_score_mean")
            pivot_tis.to_excel(xw, sheet_name="Temporal_Score")

            # Quick cross-dataset pivots for the figures.
            df.pivot_table(index="dataset", columns="model", values="fps_mean").to_excel(xw, sheet_name="FPS")
            df.pivot_table(index="dataset", columns="model", values="ssim_accuracy_mean").to_excel(xw, sheet_name="SSIM")

        _autofit(path)
        return path

    def finalize(self):
        csv = self.save_csv()
        xlsx = self.save_excel()
        print(f"\n[results_manager] wrote:\n  {csv}\n  {xlsx}")
        return csv, xlsx


def _isnan(x):
    try:
        return np.isnan(x)
    except (TypeError, ValueError):
        return False


def _autofit(xlsx_path):
    """Best-effort column auto-fit + header bold/freeze."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = load_workbook(xlsx_path)
        head_fill = PatternFill("solid", fgColor="1F4E79")
        head_font = Font(bold=True, color="FFFFFF")
        for ws in wb.worksheets:
            ws.freeze_panes = "B2"
            for cell in ws[1]:
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)
        wb.save(xlsx_path)
    except Exception as e:
        print(f"[results_manager] autofit skipped: {e}")
