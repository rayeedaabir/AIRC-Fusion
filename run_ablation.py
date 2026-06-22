"""
run_ablation.py  (UPDATED 21 Jun 2026 - robust Excel writer + per-pane LETTER labels)
-------------------------------------------------------------------------------------
Runs the proposed Vectorized FCDFusion with components switched off, and with the gamma
hyperparameter swept, on a SINGLE static pair (the paper used RGB-NIR). Writes the tables
(xlsx + CSV) and the two labelled visual panels (FIG9 components, FIG10 gamma). Each pane
of the panels carries a Cureus letter label "(A) <name>"; the printed legend maps them.

Proposed gain:  a = (ir/255)^gamma_ir ;  target = (Vmax+255)/2 [dynamic] ;
                k = a * target / Vmax + base_gain ;  fused = clip(k * vis, 0, 255)

Run:
  python run_ablation.py --vis path/rgbnir_vis.png --ir path/rgbnir_ir.png --out_dir ablation_out
"""
import os, argparse, string
import cv2, numpy as np, pandas as pd
from evaluation_metrics import compute_entropy_metric, compute_std_dev_metric, compute_ssim_accuracy_metric


def _letter(i):
    s = ""; i += 1
    while i > 0:
        i, r = divmod(i - 1, 26); s = string.ascii_uppercase[r] + s
    return s


def fuse(vis_bgr, ir_bgr, gamma_ir=2.0, base_gain=0.5, dynamic=True, gamma_vis=1.0):
    vis = vis_bgr.astype(np.float32); ir = ir_bgr.astype(np.float32)
    Vmax = np.maximum.reduce([vis[:, :, 0], vis[:, :, 1], vis[:, :, 2]]); Vmax = np.maximum(Vmax, 1.0)
    if gamma_vis != 1.0:
        Vmax = 255.0 * (Vmax / 255.0) ** gamma_vis; Vmax = np.maximum(Vmax, 1.0)
    a = (ir[:, :, 0] / 255.0) ** gamma_ir
    target = (Vmax + 255.0) / 2.0 if dynamic else (255.0 + 255.0) / 2.0
    k = a * target / Vmax + base_gain
    return np.clip(vis * k[:, :, None], 0, 255).astype(np.uint8)


def metrics(f, v, ir):
    return dict(SSIM=round(float(compute_ssim_accuracy_metric(f, v, ir)), 4),
                Entropy=round(float(compute_entropy_metric(f)), 4),
                StdDev=round(float(compute_std_dev_metric(f)), 4))


def label_tile(img, txt):
    h, w = img.shape[:2]; tw = 300; img = cv2.resize(img, (tw, int(h * tw / w)))
    bar = np.full((26, tw, 3), 30, np.uint8)
    cv2.putText(bar, txt[:42], (5, 18), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1, cv2.LINE_AA)
    return np.vstack([bar, img])


def lettered_tiles(items):
    """items = list of (img, name); returns (tiles, legend_string)."""
    tiles, legend = [], []
    for i, (img, name) in enumerate(items):
        let = _letter(i)
        tiles.append(label_tile(img, f"({let}) {name}"))
        legend.append(f"({let}) {name}")
    return tiles, "; ".join(legend)


def grid(tiles, cols=4):
    h = max(t.shape[0] for t in tiles); w = max(t.shape[1] for t in tiles)
    tiles = [np.vstack([t, np.zeros((h - t.shape[0], t.shape[1], 3), np.uint8)]) for t in tiles]
    rows = [np.hstack(tiles[i:i + cols] + [np.zeros((h, w, 3), np.uint8)] * ((cols - len(tiles[i:i + cols])) % cols))
            for i in range(0, len(tiles), cols)]
    return np.vstack(rows)


def write_xlsx(path, sheets):
    """Robust .xlsx writer (never trips 'At least one sheet must be visible'); also drops CSVs."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook(); wb.remove(wb.active)
    for name, df in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        df.to_csv(os.path.join(os.path.dirname(path), f"{name}.csv"), index=False)
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    wb.active = 0
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--vis", required=True); ap.add_argument("--ir", required=True)
    ap.add_argument("--out_dir", default="ablation_out"); a = ap.parse_args()
    os.makedirs(a.out_dir, exist_ok=True)
    vis = cv2.imread(a.vis); ir = cv2.imread(a.ir)
    if vis is None or ir is None:
        raise SystemExit("could not read --vis/--ir")
    if ir.shape[:2] != vis.shape[:2]:
        ir = cv2.resize(ir, (vis.shape[1], vis.shape[0]))

    # ---- Table 8 / FIG 9: component ablation ----
    comp = [("Full FCDFusion (proposed)", dict()),
            ("No gamma (gamma=1)", dict(gamma_ir=1.0)),
            ("No base gain (+0.5 removed)", dict(base_gain=0.0)),
            ("No dynamic target", dict(dynamic=False)),
            ("No gamma & no base gain", dict(gamma_ir=1.0, base_gain=0.0)),
            ("No gamma & no dynamic target", dict(gamma_ir=1.0, dynamic=False)),
            ("No base gain & no dynamic target", dict(base_gain=0.0, dynamic=False)),
            ("All components removed", dict(gamma_ir=1.0, base_gain=0.0, dynamic=False))]
    rows8 = []; items8 = [(vis, "Visible"), (ir, "Infrared")]
    for name, kw in comp:
        f = fuse(vis, ir, **kw); rows8.append({"Variant": name, **metrics(f, vis, ir)}); items8.append((f, name))
    df8 = pd.DataFrame(rows8)
    tiles8, legend8 = lettered_tiles(items8)

    # ---- Table 9 / FIG 10: gamma tuning ----
    gam = [("gamma=1 on IR", dict(gamma_ir=1.0)), ("gamma=1.5 on IR", dict(gamma_ir=1.5)),
           ("gamma=2 on IR (proposed)", dict(gamma_ir=2.0)), ("gamma=3 on IR", dict(gamma_ir=3.0)),
           ("gamma=4 on IR", dict(gamma_ir=4.0)), ("gamma=2 on Vis", dict(gamma_ir=1.0, gamma_vis=2.0)),
           ("gamma=2 on Vis & IR", dict(gamma_ir=2.0, gamma_vis=2.0))]
    rows9 = []; items9 = [(vis, "Visible"), (ir, "Infrared")]
    for name, kw in gam:
        f = fuse(vis, ir, **kw); rows9.append({"Setting": name, **metrics(f, vis, ir)}); items9.append((f, name))
    df9 = pd.DataFrame(rows9)
    tiles9, legend9 = lettered_tiles(items9)

    xl = os.path.join(a.out_dir, "ablation_tables.xlsx")
    write_xlsx(xl, {"Table8_components": df8, "Table9_gamma": df9})
    cv2.imwrite(os.path.join(a.out_dir, "FIG9_ablation.png"), grid(tiles8))
    cv2.imwrite(os.path.join(a.out_dir, "FIG10_gamma.png"), grid(tiles9))
    print("Table 8 (components):\n", df8.to_string(index=False))
    print("\nTable 9 (gamma):\n", df9.to_string(index=False))
    print(f"\nFIG9 legend: {legend8}.")
    print(f"FIG10 legend: {legend9}.")
    print(f"\nWrote {xl} (+ CSV copies), FIG9_ablation.png, FIG10_gamma.png in {a.out_dir}")


if __name__ == "__main__":
    main()
# --- end of file (sentinel) ---